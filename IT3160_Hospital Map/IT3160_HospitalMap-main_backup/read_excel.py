import sys, json
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

def get_cell_color_type(cell):
    """Map fill color to room type."""
    fill = cell.fill
    if not fill or fill.patternType in (None, 'none'):
        return 'path'
    fg = fill.fgColor
    if not fg:
        return 'path'
    
    color = None
    if fg.type == 'rgb':
        color = fg.rgb.upper()
    elif fg.type == 'theme':
        # theme:6 is orange (waiting area)
        return 'room-normal'  # treat as room
    else:
        return 'path'
    
    if not color or color in ('00000000', 'FFFFFFFF', 'FF000000'):
        return 'path'
    
    # Green = nhà vệ sinh (WC)
    if color in ('FF00FF00',):
        return 'room-wc'
    # Light green = stairs/elevator
    if color in ('FFB6D7A8', 'FFD9EAD3'):
        return 'elevator'
    # Gray = gate/wall boundary
    if color in ('FFADADAD', 'FFBFBFBF', 'FF9E9E9E'):
        return 'gate'
    # Yellow = normal room
    if 'FF' in color and color != 'FFFFFFFF':
        return 'room-normal'
    return 'path'

def read_floor_full(filename, floor_name):
    wb = openpyxl.load_workbook(filename, data_only=True)
    ws = wb.active
    
    # Get merged cell ranges - map each cell to its merged cell value
    merged_map = {}
    for merged_range in ws.merged_cells.ranges:
        min_row = merged_range.min_row
        min_col = merged_range.min_col
        max_row = merged_range.max_row
        max_col = merged_range.max_col
        # Get value from top-left
        top_left = ws.cell(min_row, min_col)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                merged_map[(r, c)] = {
                    'value': top_left.value,
                    'is_origin': (r == min_row and c == min_col),
                    'rowspan': max_row - min_row + 1,
                    'colspan': max_col - min_col + 1,
                    'origin_row': min_row,
                    'origin_col': min_col,
                }
    
    MAX_ROW = ws.max_row
    MAX_COL = ws.max_column
    
    # Only consider rows 1-21, cols 1-8 (based on visible data)
    USE_ROWS = min(MAX_ROW, 21)
    USE_COLS = min(MAX_COL, 8)
    
    print(f"\n{'='*60}")
    print(f"Floor: {floor_name}  ({USE_ROWS} rows x {USE_COLS} cols)")
    print(f"{'='*60}")
    
    grid = []
    for r in range(1, USE_ROWS + 1):
        row_data = []
        for c in range(1, USE_COLS + 1):
            cell = ws.cell(r, c)
            
            if (r, c) in merged_map:
                info = merged_map[(r, c)]
                val = str(info['value']).strip() if info['value'] else ''
                is_origin = info['is_origin']
            else:
                val = str(cell.value).strip() if cell.value else ''
                is_origin = True
            
            color_type = get_cell_color_type(cell)
            
            row_data.append({
                'r': r, 'c': c,
                'value': val,
                'type': color_type,
                'is_origin': is_origin,
                'in_merged': (r, c) in merged_map,
            })
        grid.append(row_data)
    
    # Print grid summary
    for r_idx, row in enumerate(grid):
        row_str = f"R{r_idx+1:02d}: "
        for cell in row:
            if cell['value']:
                v = cell['value'][:12].ljust(12)
                row_str += f"[{v}]"
            elif cell['type'] in ('room-normal', 'room-wc', 'elevator', 'gate'):
                row_str += f"[{'·':12}]"
            else:
                row_str += f"[{'':12}]"
        print(row_str)
    
    return grid

f1 = read_floor_full("Bản đồ bệnh viện_tầng 1.xlsx", "Tầng 1")
f2 = read_floor_full("Bản đồ bệnh viện_tầng 2.xlsx", "Tầng 2")
